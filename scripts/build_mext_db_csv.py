#!/usr/bin/env python3
"""Build DB-oriented CSV files from downloaded MEXT Excel workbooks.

Outputs:
- institutions.csv
- faculties.csv
- departments.csv
- excluded_sources.csv

The CSVs are shaped for the current institution/faculty/department schema.
They keep source metadata columns so import code can trace each row back to the
MEXT workbook and worksheet.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
from collections import OrderedDict
from concurrent.futures import ProcessPoolExecutor
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

# scripts/ ディレクトリ内の兄弟モジュール (実行時 sys.path[0] = scripts/)
from academic_field import classify_field, classify_track
from mext_common import (
    DISPLAY_LANGUAGE,
    REGION_CODE,
    DepartmentRow,
    FacultyRow,
    InstitutionRow,
    normalize_text,
    public_id,
)

try:
    from openpyxl import load_workbook
except ModuleNotFoundError as exc:  # pragma: no cover - environment guard
    raise SystemExit(
        "openpyxl is required. Install it with `python3 -m pip install openpyxl` "
        "or run this script with the bundled Codex Python runtime."
    ) from exc


DEFAULT_INPUT_MANIFEST = Path("data/raw/mext_excels/manifest.csv")
DEFAULT_OUTPUT_DIR = Path("data/processed/mext_db_csv")
INSTITUTION_TYPE_ENUM = {
    "university",
    "graduate_school",
    "junior_college",
    "technical_college",
    "technical_college_advanced",
    "high_school",
    "vocational_school",
}

PREFECTURE_CODES = {
    "北海道": "JP-01",
    "青森県": "JP-02",
    "岩手県": "JP-03",
    "宮城県": "JP-04",
    "秋田県": "JP-05",
    "山形県": "JP-06",
    "福島県": "JP-07",
    "茨城県": "JP-08",
    "栃木県": "JP-09",
    "群馬県": "JP-10",
    "埼玉県": "JP-11",
    "千葉県": "JP-12",
    "東京都": "JP-13",
    "神奈川県": "JP-14",
    "新潟県": "JP-15",
    "富山県": "JP-16",
    "石川県": "JP-17",
    "福井県": "JP-18",
    "山梨県": "JP-19",
    "長野県": "JP-20",
    "岐阜県": "JP-21",
    "静岡県": "JP-22",
    "愛知県": "JP-23",
    "三重県": "JP-24",
    "滋賀県": "JP-25",
    "京都府": "JP-26",
    "大阪府": "JP-27",
    "兵庫県": "JP-28",
    "奈良県": "JP-29",
    "和歌山県": "JP-30",
    "鳥取県": "JP-31",
    "島根県": "JP-32",
    "岡山県": "JP-33",
    "広島県": "JP-34",
    "山口県": "JP-35",
    "徳島県": "JP-36",
    "香川県": "JP-37",
    "愛媛県": "JP-38",
    "高知県": "JP-39",
    "福岡県": "JP-40",
    "佐賀県": "JP-41",
    "長崎県": "JP-42",
    "熊本県": "JP-43",
    "大分県": "JP-44",
    "宮崎県": "JP-45",
    "鹿児島県": "JP-46",
    "沖縄県": "JP-47",
}

IN_SCOPE_SEQUENCES = {
    "university": {"01", "02", "05"},
    "junior_college": {"01"},
    "kosen": {"01"},
}


@dataclass
class ExcludedSourceRow:
    source_index: str
    category: str
    fiscal_year_label: str
    file_label: str
    file_sequence: str
    output_path: str
    reason: str


def cell(row: list[Any], one_based_index: int) -> str:
    index = one_based_index - 1
    if index >= len(row):
        return ""
    return normalize_text(row[index])


def parse_int(value: str | None) -> int | None:
    if value is None or value == "":
        return None
    return int(value)


def load_manifest(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8", newline="") as file:
        return list(csv.DictReader(file))


def name_pair(display_name: str, english_name: str = "", suffix_en: str = "", suffix_ja: str = "") -> tuple[str, str, str]:
    display = f"{display_name}{suffix_ja}"
    if english_name:
        return f"{english_name}{suffix_en}", display, "source_english"
    return display, display, "fallback_display_name"


def graduate_suffixes(display_name: str, english_name: str) -> tuple[str, str]:
    if "大学院大学" in display_name or display_name.endswith("大学院"):
        return "", ""
    if english_name and "graduate" in english_name.lower():
        return "", "大学院"
    return (" Graduate School" if english_name else "", "大学院")


def institution_suffixes(institution_type: str, display_name: str, english_name: str) -> tuple[str, str]:
    if institution_type == "graduate_school":
        return graduate_suffixes(display_name, english_name)
    if institution_type == "technical_college_advanced":
        suffix_ja = "" if display_name.endswith("専攻科") else "専攻科"
        suffix_en = "" if not english_name or "advanced" in english_name.lower() else " Advanced Course"
        return suffix_en, suffix_ja
    return "", ""


def parse_title(value: str) -> tuple[str, str, str]:
    text = normalize_text(value)
    match = re.match(r"^(国立|公立|私立)\s+(.+?)(?:（(.+)）)?$", text)
    if match:
        return match.group(1), normalize_text(match.group(2)), normalize_text(match.group(3))
    match = re.match(r"^(.+?)(?:（(.+)）)?$", text)
    if match:
        return "", normalize_text(match.group(1)), normalize_text(match.group(2))
    return "", text, ""


def rows_from_sheet(worksheet: Any) -> list[list[Any]]:
    return [list(row) for row in worksheet.iter_rows(values_only=True)]


def non_empty_values(row: list[Any]) -> list[str]:
    return [normalize_text(value) for value in row if normalize_text(value)]


def is_single_heading(row: list[Any]) -> bool:
    values = non_empty_values(row)
    return len(values) == 1 and bool(cell(row, 2))


def find_section(rows: list[list[Any]], heading: str) -> int | None:
    for index, row in enumerate(rows):
        if cell(row, 2) == heading and is_single_heading(row):
            return index
    return None


def section_rows(rows: list[list[Any]], heading: str) -> list[tuple[int, list[Any]]]:
    start = find_section(rows, heading)
    if start is None:
        return []

    result: list[tuple[int, list[Any]]] = []
    seen_data = False
    blank_count = 0
    for row_number, row in enumerate(rows[start + 2 :], start=start + 3):
        if is_single_heading(row):
            break
        values = non_empty_values(row)
        if not values:
            if seen_data:
                blank_count += 1
                if blank_count >= 10:
                    break
            continue
        blank_count = 0
        if cell(row, 2) == heading:
            continue
        if cell(row, 2) in {"－", "-", "ー"} and len(values) == 1:
            continue
        result.append((row_number, row))
        seen_data = True
    return result


def extract_address(rows: list[list[Any]], category: str) -> str:
    basic_address = cell(rows[4], 12) if len(rows) >= 5 else ""
    if category in {"kosen", "junior_college"}:
        for _row_number, row in section_rows(rows, "学科所在地（キャンパス名など）"):
            campus_address = cell(row, 9)
            if campus_address:
                return campus_address
    return basic_address


# キャンパスセクションのヘッダ (大学/院 と 高専/短大 で名称が異なる)
CAMPUS_SECTION_HEADERS = (
    "学部・研究科所在地（キャンパス名など）",
    "学科所在地（キャンパス名など）",
)


def campus_prefectures(rows: list[list[Any]]) -> set[str]:
    """キャンパスセクションから、全キャンパスの都道府県コード集合を返す。

    institutions は単一住所しか持てないため、複数キャンパス (例: 立命館=京都/滋賀/大阪)
    の都道府県は検索レイヤ (institution_prefectures) 側で保持する。その元データ。
    """
    prefs: set[str] = set()
    for header in CAMPUS_SECTION_HEADERS:
        for _row_number, row in section_rows(rows, header):
            code = prefecture_code(cell(row, 9))
            if code:
                prefs.add(code)
    return prefs


def prefecture_code(name_or_address: str) -> str:
    value = normalize_text(name_or_address)
    if value in PREFECTURE_CODES:
        return PREFECTURE_CODES[value]
    for name, code in PREFECTURE_CODES.items():
        if value.startswith(name):
            return code
    return ""


def source_is_in_scope(row: dict[str, str]) -> tuple[bool, str]:
    category = row["category"]
    sequence = row["file_sequence"]
    if category == "university":
        if sequence in IN_SCOPE_SEQUENCES["university"] or sequence.startswith("03-"):
            return True, ""
        return False, "not_an_institution_workbook_for_current_schema"
    if category == "junior_college":
        if sequence in IN_SCOPE_SEQUENCES["junior_college"] or sequence.startswith("02-"):
            return True, ""
        return False, "not_an_institution_workbook_for_current_schema"
    if category == "kosen":
        if sequence in IN_SCOPE_SEQUENCES["kosen"]:
            return True, ""
        return False, "kosen_supplement_not_in_current_schema"
    return False, "no_target_table_in_current_schema"


def latest_year(rows: list[dict[str, str]]) -> int | None:
    years = [parse_int(row.get("fiscal_year_ad")) for row in rows]
    years = [year for year in years if year is not None]
    return max(years) if years else None


def add_unique(ordered: OrderedDict[str, Any], key: str, value: Any) -> Any:
    if key not in ordered:
        ordered[key] = value
    return ordered[key]


def add_institution(
    institutions: OrderedDict[str, InstitutionRow],
    row: dict[str, str],
    sheet_name: str,
    source_title: str,
    display_name: str,
    english_name: str,
    source_prefix: str,
    institution_type: str,
    address: str,
) -> InstitutionRow:
    suffix_en, suffix_ja = institution_suffixes(institution_type, display_name, english_name)
    name, display, name_source = name_pair(
        display_name=display_name,
        english_name=english_name,
        suffix_en=suffix_en,
        suffix_ja=suffix_ja,
    )
    key = f"{institution_type}|{display}|{REGION_CODE}"
    pid = public_id("institution", key)
    institution = InstitutionRow(
        public_id=pid,
        name=name,
        display_name=display,
        display_language=DISPLAY_LANGUAGE,
        region_code=REGION_CODE,
        institution_type=institution_type,
        prefecture_code=prefecture_code(address),
        address=address,
        website_url="",
        deleted_at="",
        source_fiscal_year_label=row["fiscal_year_label"],
        source_fiscal_year_ad=row["fiscal_year_ad"],
        source_category=row["category"],
        source_file_label=row["file_label"],
        source_workbook_path=row["output_path"],
        source_sheet_name=sheet_name,
        source_title=source_title,
        source_name_prefix=source_prefix,
        source_name_english=english_name,
        name_source=name_source,
    )
    return add_unique(institutions, key, institution)


def add_faculty(
    faculties: OrderedDict[str, FacultyRow],
    institution: InstitutionRow,
    row: dict[str, str],
    sheet_name: str,
    display_name: str,
    source_section: str,
    source_row: int,
) -> FacultyRow:
    name, display, name_source = name_pair(display_name=display_name)
    key = f"{institution.public_id}|{display}"
    pid = public_id("faculty", key)
    faculty = FacultyRow(
        public_id=pid,
        institution_public_id=institution.public_id,
        name=name,
        display_name=display,
        display_language=DISPLAY_LANGUAGE,
        region_code=REGION_CODE,
        academic_field=classify_field(display),
        deleted_at="",
        source_section=source_section,
        source_row=source_row,
        source_fiscal_year_label=row["fiscal_year_label"],
        source_category=row["category"],
        source_file_label=row["file_label"],
        source_workbook_path=row["output_path"],
        source_sheet_name=sheet_name,
        name_source=name_source,
    )
    return add_unique(faculties, key, faculty)


def add_department(
    departments: OrderedDict[str, DepartmentRow],
    faculty: FacultyRow,
    row: dict[str, str],
    sheet_name: str,
    display_name: str,
    source_section: str,
    source_row: int,
    source_prefecture_name: str,
    source_city: str,
) -> DepartmentRow:
    name, display, name_source = name_pair(display_name=display_name)
    key = f"{faculty.public_id}|{display}"
    pid = public_id("department", key)
    # 学科名で分類し、不明なら学部の分野を継承する
    field = classify_field(display) or faculty.academic_field
    track = classify_track(field, display, faculty.display_name)
    department = DepartmentRow(
        public_id=pid,
        faculty_public_id=faculty.public_id,
        name=name,
        display_name=display,
        display_language=DISPLAY_LANGUAGE,
        region_code=REGION_CODE,
        academic_field=field,
        academic_track=track,
        deleted_at="",
        source_section=source_section,
        source_row=source_row,
        source_prefecture_code=prefecture_code(source_prefecture_name),
        source_prefecture_name=source_prefecture_name,
        source_city=source_city,
        source_fiscal_year_label=row["fiscal_year_label"],
        source_category=row["category"],
        source_file_label=row["file_label"],
        source_workbook_path=row["output_path"],
        source_sheet_name=sheet_name,
        name_source=name_source,
    )
    return add_unique(departments, key, department)


def department_display(primary: str, secondary: str) -> str:
    primary = normalize_text(primary)
    secondary = normalize_text(secondary)
    if primary and secondary:
        return f"{primary} {secondary}"
    return primary or secondary


def process_school_sheet(
    manifest_row: dict[str, str],
    worksheet: Any,
    institutions: OrderedDict[str, InstitutionRow],
    faculties: OrderedDict[str, FacultyRow],
    departments: OrderedDict[str, DepartmentRow],
    campus_prefs: dict[str, set[str]],
) -> None:
    rows = rows_from_sheet(worksheet)
    source_title = cell(rows[0], 2) if rows else worksheet.title
    source_prefix, display_name, english_name = parse_title(source_title)
    if not display_name:
        display_name = worksheet.title
    category = manifest_row["category"]
    address = extract_address(rows, category)
    sheet_campus_prefs = campus_prefectures(rows)

    def record_campuses(inst: InstitutionRow | None) -> None:
        if inst is None:
            return
        prefs = campus_prefs.setdefault(inst.public_id, set())
        prefs.update(sheet_campus_prefs)
        if inst.prefecture_code:  # 本部の県も必ず含める
            prefs.add(inst.prefecture_code)

    if category == "university":
        undergrad_rows = section_rows(rows, "学部")
        graduate_rows = section_rows(rows, "研究科")
        has_undergrad = any(cell(item, 2) and cell(item, 4) for _row_number, item in undergrad_rows)
        has_graduate = any(cell(item, 2) and cell(item, 4) for _row_number, item in graduate_rows)

        university: InstitutionRow | None = None
        if has_undergrad:
            university = add_institution(
                institutions,
                manifest_row,
                worksheet.title,
                source_title,
                display_name,
                english_name,
                source_prefix,
                "university",
                address,
            )
            record_campuses(university)
            for row_number, item in undergrad_rows:
                faculty_name = cell(item, 2)
                department_name = cell(item, 4)
                if not faculty_name or not department_name:
                    continue
                faculty = add_faculty(faculties, university, manifest_row, worksheet.title, faculty_name, "学部", row_number)
                add_department(
                    departments,
                    faculty,
                    manifest_row,
                    worksheet.title,
                    department_name,
                    "学部",
                    row_number,
                    cell(item, 7),
                    cell(item, 9),
                )

        if has_graduate:
            graduate = add_institution(
                institutions,
                manifest_row,
                worksheet.title,
                source_title,
                display_name,
                english_name,
                source_prefix,
                "graduate_school",
                address,
            )
            record_campuses(graduate)
            for row_number, item in graduate_rows:
                faculty_name = cell(item, 2)
                department_name = cell(item, 4)
                if not faculty_name or not department_name:
                    continue
                faculty = add_faculty(faculties, graduate, manifest_row, worksheet.title, faculty_name, "研究科", row_number)
                add_department(
                    departments,
                    faculty,
                    manifest_row,
                    worksheet.title,
                    department_name,
                    "研究科",
                    row_number,
                    cell(item, 7),
                    cell(item, 9),
                )

        target = university
        if target is not None:
            process_special_course(manifest_row, worksheet.title, rows, target, faculties, departments)
        return

    if category == "junior_college":
        institution = add_institution(
            institutions,
            manifest_row,
            worksheet.title,
            source_title,
            display_name,
            english_name,
            source_prefix,
            "junior_college",
            address,
        )
        record_campuses(institution)
        synthetic = add_faculty(faculties, institution, manifest_row, worksheet.title, "短期大学", "学科", 0)
        for row_number, item in section_rows(rows, "学科"):
            display = department_display(cell(item, 2), cell(item, 4))
            if not display:
                continue
            add_department(
                departments,
                synthetic,
                manifest_row,
                worksheet.title,
                display,
                "学科",
                row_number,
                cell(item, 7),
                cell(item, 9),
            )
        process_special_course(manifest_row, worksheet.title, rows, institution, faculties, departments)
        return

    if category == "kosen":
        institution = add_institution(
            institutions,
            manifest_row,
            worksheet.title,
            source_title,
            display_name,
            english_name,
            source_prefix,
            "technical_college",
            address,
        )
        record_campuses(institution)
        synthetic = add_faculty(faculties, institution, manifest_row, worksheet.title, "本科", "学科", 0)
        for row_number, item in section_rows(rows, "学科"):
            display = department_display(cell(item, 2), cell(item, 4))
            if not display:
                continue
            add_department(
                departments,
                synthetic,
                manifest_row,
                worksheet.title,
                display,
                "学科",
                row_number,
                "",
                "",
            )
        if section_rows(rows, "専攻科"):
            advanced = add_institution(
                institutions,
                manifest_row,
                worksheet.title,
                source_title,
                display_name,
                english_name,
                source_prefix,
                "technical_college_advanced",
                address,
            )
            record_campuses(advanced)
            process_special_course(manifest_row, worksheet.title, rows, advanced, faculties, departments)


def process_special_course(
    manifest_row: dict[str, str],
    sheet_name: str,
    rows: list[list[Any]],
    institution: InstitutionRow,
    faculties: OrderedDict[str, FacultyRow],
    departments: OrderedDict[str, DepartmentRow],
) -> None:
    special_course_rows = section_rows(rows, "専攻科")
    if not special_course_rows:
        return
    faculty = add_faculty(faculties, institution, manifest_row, sheet_name, "専攻科", "専攻科", 0)
    for row_number, item in special_course_rows:
        display = department_display(cell(item, 2), cell(item, 4))
        if not display:
            continue
        add_department(
            departments,
            faculty,
            manifest_row,
            sheet_name,
            display,
            "専攻科",
            row_number,
            cell(item, 7),
            cell(item, 9),
        )


def validate_institution_types(rows: list[InstitutionRow]) -> None:
    invalid = sorted({row.institution_type for row in rows if row.institution_type not in INSTITUTION_TYPE_ENUM})
    if invalid:
        raise RuntimeError(f"unknown institution_type values: {', '.join(invalid)}")


def process_manifest_row(
    row: dict[str, str],
    institutions: OrderedDict[str, InstitutionRow],
    faculties: OrderedDict[str, FacultyRow],
    departments: OrderedDict[str, DepartmentRow],
    campus_prefs: dict[str, set[str]],
) -> None:
    workbook = load_workbook(row["output_path"], read_only=True, data_only=True)
    try:
        for worksheet in workbook.worksheets:
            process_school_sheet(row, worksheet, institutions, faculties, departments, campus_prefs)
    finally:
        workbook.close()


WorkbookResult = tuple[
    "OrderedDict[str, InstitutionRow]",
    "OrderedDict[str, FacultyRow]",
    "OrderedDict[str, DepartmentRow]",
    dict[str, set[str]],
]


def process_workbook(row: dict[str, str]) -> WorkbookResult:
    """1 workbook を独立に解析し、ローカルな ordered dict を返す (並列ワーカ)。

    workbook 単位で完結する純関数なので multiprocessing で安全に並列化できる。
    workbook 内の重複排除 (add_unique = 先勝ち) はここで、workbook 間の重複排除は
    呼び出し側が scoped 順に再 add_unique することで担保する。読みやキャンパス県は
    入力から決定的に算出されるため、どのプロセスで計算しても結果は同一になる。
    """
    institutions: OrderedDict[str, InstitutionRow] = OrderedDict()
    faculties: OrderedDict[str, FacultyRow] = OrderedDict()
    departments: OrderedDict[str, DepartmentRow] = OrderedDict()
    campus_prefs: dict[str, set[str]] = {}
    process_manifest_row(row, institutions, faculties, departments, campus_prefs)
    return institutions, faculties, departments, campus_prefs


def merge_workbook_result(
    result: WorkbookResult,
    institutions: OrderedDict[str, InstitutionRow],
    faculties: OrderedDict[str, FacultyRow],
    departments: OrderedDict[str, DepartmentRow],
    campus_prefs: dict[str, set[str]],
) -> None:
    inst, fac, dep, camp = result
    for key, value in inst.items():
        add_unique(institutions, key, value)
    for key, value in fac.items():
        add_unique(faculties, key, value)
    for key, value in dep.items():
        add_unique(departments, key, value)
    for pid, codes in camp.items():
        campus_prefs.setdefault(pid, set()).update(codes)


def write_rows(path: Path, rows: list[Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = list(asdict(rows[0]).keys()) if rows else []
    with path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow(asdict(row))


def write_prefecture_rows(path: Path, rows: list[tuple[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(["institution_public_id", "prefecture_code"])
        writer.writerows(rows)


def check_duplicate_public_ids(rows: list[Any], label: str) -> None:
    seen: set[str] = set()
    for row in rows:
        if row.public_id in seen:
            raise RuntimeError(f"duplicate public_id in {label}: {row.public_id}")
        seen.add(row.public_id)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input-manifest", type=Path, default=DEFAULT_INPUT_MANIFEST)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--latest-only", action="store_true", default=True, help="Use only the latest fiscal year in the manifest. Default: true.")
    parser.add_argument("--all-years", action="store_true", help="Process all fiscal years in the manifest.")
    parser.add_argument("--fiscal-year", type=int, default=None, help="Process one fiscal year. Accepts Reiwa year or western year.")
    parser.add_argument(
        "--jobs",
        type=int,
        default=0,
        help="Number of parallel worker processes for workbook parsing. 0 (default) = number of CPUs. 1 = sequential.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    manifest_rows = load_manifest(args.input_manifest)
    if args.all_years:
        args.latest_only = False
    if args.fiscal_year is not None:
        if args.fiscal_year >= 1900:
            manifest_rows = [row for row in manifest_rows if parse_int(row.get("fiscal_year_ad")) == args.fiscal_year]
        else:
            manifest_rows = [row for row in manifest_rows if parse_int(row.get("fiscal_year_reiwa")) == args.fiscal_year]
    elif args.latest_only:
        latest = latest_year(manifest_rows)
        manifest_rows = [row for row in manifest_rows if parse_int(row.get("fiscal_year_ad")) == latest]

    institutions: OrderedDict[str, InstitutionRow] = OrderedDict()
    faculties: OrderedDict[str, FacultyRow] = OrderedDict()
    departments: OrderedDict[str, DepartmentRow] = OrderedDict()
    campus_prefs: dict[str, set[str]] = {}
    excluded: list[ExcludedSourceRow] = []

    scoped_rows: list[dict[str, str]] = []
    for row in manifest_rows:
        in_scope, reason = source_is_in_scope(row)
        if in_scope:
            scoped_rows.append(row)
        else:
            excluded.append(
                ExcludedSourceRow(
                    source_index=row["index"],
                    category=row["category"],
                    fiscal_year_label=row["fiscal_year_label"],
                    file_label=row["file_label"],
                    file_sequence=row["file_sequence"],
                    output_path=row["output_path"],
                    reason=reason,
                )
            )

    def progress(index: int, row: dict[str, str]) -> None:
        print(f"[{index}/{len(scoped_rows)}] {row['category']}/{row['fiscal_year_label']}: {row['file_label']}", flush=True)

    jobs = args.jobs if args.jobs > 0 else (os.cpu_count() or 1)
    jobs = max(1, min(jobs, len(scoped_rows)))
    if jobs == 1:
        for index, row in enumerate(scoped_rows, start=1):
            progress(index, row)
            process_manifest_row(row, institutions, faculties, departments, campus_prefs)
    else:
        # workbook 単位で並列パース。ProcessPoolExecutor.map は入力順で結果を返すため、
        # scoped 順での再 add_unique = 逐次実行と同じ「先勝ち」順になり出力は決定的。
        with ProcessPoolExecutor(max_workers=jobs) as executor:
            for index, (row, result) in enumerate(
                zip(scoped_rows, executor.map(process_workbook, scoped_rows)), start=1
            ):
                progress(index, row)
                merge_workbook_result(result, institutions, faculties, departments, campus_prefs)

    institution_rows = list(institutions.values())
    faculty_rows = list(faculties.values())
    department_rows = list(departments.values())

    check_duplicate_public_ids(institution_rows, "institutions")
    check_duplicate_public_ids(faculty_rows, "faculties")
    check_duplicate_public_ids(department_rows, "departments")
    validate_institution_types(institution_rows)

    args.output_dir.mkdir(parents=True, exist_ok=True)
    write_rows(args.output_dir / "institutions.csv", institution_rows)
    write_rows(args.output_dir / "faculties.csv", faculty_rows)
    write_rows(args.output_dir / "departments.csv", department_rows)
    write_rows(args.output_dir / "excluded_sources.csv", excluded)

    # 検索レイヤ用: 機関ごとの全キャンパス都道府県 (本部 + 各キャンパス)。
    # institutions は単一住所なので、複数県にまたがる大学はここで全県を保持する。
    prefecture_rows = sorted(
        ((pid, code) for pid, codes in campus_prefs.items() for code in codes),
        key=lambda x: (x[0], x[1]),
    )
    write_prefecture_rows(args.output_dir / "institution_prefectures.csv", prefecture_rows)

    multi_campus = sum(1 for codes in campus_prefs.values() if len(codes) > 1)
    summary = {
        "source_rows": len(manifest_rows),
        "processed_workbooks": len(scoped_rows),
        "excluded_sources": len(excluded),
        "institutions": len(institution_rows),
        "faculties": len(faculty_rows),
        "departments": len(department_rows),
        "institution_prefecture_rows": len(prefecture_rows),
        "multi_prefecture_institutions": multi_campus,
        "output_dir": str(args.output_dir),
    }
    with (args.output_dir / "summary.json").open("w", encoding="utf-8") as file:
        json.dump(summary, file, ensure_ascii=False, indent=2)
        file.write("\n")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
